import regex as re
import pandas as pd
from icalendar import Event, Calendar
from datetime import datetime, timedelta
import openpyxl
import os
from typing import List


def _get_time_row(df: pd.DataFrame) -> pd.Series:
    """
    Get the time row from the dataframe.

    Parameters
    ----------
    df : pandas.DataFrame
        The dataframe to get the time row from.

    Returns
    -------
    pandas.Series
        The time row from the dataframe.
    """
    for row in df.iterrows():
        if re.match(r"^\d{1,2}:\d{1,2}-\d{1,2}:\d{1,2}$", str(row[1].iloc[1])):
            return row


def _get_daily_table(df: pd.DataFrame, class_pattern: str) -> pd.DataFrame:
    """Get the simplified dataframe for a given class."""
    df = df.copy()
    
    time_row = _get_time_row(df)
    new_cols = time_row[1].to_list()
    new_cols.pop(0)
    new_cols.insert(0, "Classroom")
    df.columns = new_cols
    
    df.set_index("Classroom", inplace=True)
    df = df.iloc[time_row[0] + 1 :]
    
    dept, year = class_pattern.split()
    
    patterns = [
        # basic pattern (e.g "CE 4", "CE 4A")
        fr"{dept}\s*{year}[A-Z]?",
        
        # multiple sections (e.g "CE 4A, 4B")
        fr"{dept}\s*{year}[A-Z](\s*,\s*{year}[A-Z])*",
        
        # department with sections combined (e.g "CE 4A, CE 4B")
        fr"{dept}\s*{year}[A-Z](\s*,\s*{dept}\s*{year}[A-Z])*",
        
        # course numbers starting with the year number (e.g., CE 459, CE/RN 459)
        fr"{dept}\s*{year}[0-9]{{2}}",
        
        # multiple departments sharing course number starting with year
        fr"(?:[A-Z]{{2,3}}(?:\s*[,/]\s*)?)*{dept}(?:\s*[,/]\s*[A-Z]{{2,3}})*\s+{year}[0-9]{{2}}",
        
        # department mentioned first in shared course
        fr"{dept}(?:\s*[,/]\s*[A-Z]{{2,3}})+\s+{year}[0-9]{{2}}"
    ]
    
    combined_pattern = '|'.join(f'({pattern})' for pattern in patterns)
    
    df = df.mask(~df.map(lambda x: bool(re.search(combined_pattern, str(x), re.IGNORECASE))))
    df = df.dropna(how="all")
    
    return df


def _get_all_daily_tables(filename: str, class_pattern: str) -> dict:
    """
    Get all the daily tables from an excel file.

    Parameters
    ----------
    filename : str
        The filename of the excel file to get the daily tables from.
    class_pattern : str
        The class to get the daily tables or. E.g. 'EL 3'

    Returns
    -------
    dict
        A dictionary of the daily tables for each class.
    """
    filename += ".xlsx"

    workbook = openpyxl.load_workbook(filename)
    dfs = {}
    for sheet in workbook.sheetnames:
        merged_cells = workbook[sheet].merged_cells.ranges
        for mc in merged_cells.copy():
            if mc.max_col - mc.min_col == 1:
                merged_value = workbook[sheet].cell(mc.min_row, mc.min_col).value
                workbook[sheet].unmerge_cells(mc.coord)
                workbook[sheet].cell(mc.min_row, mc.min_col).value = merged_value
                workbook[sheet].cell(mc.max_row, mc.max_col).value = merged_value

        data = workbook[sheet].values
        header = next(data)
        df = pd.DataFrame(data, columns=header)
        df = df.dropna(axis=1, how="all")

        dfs[sheet] = _get_daily_table(df, class_pattern)

    return dfs


def get_time_table(filename: str, class_pattern: str) -> pd.DataFrame:
    """
    Get the complete time table for a particular class for all days.

    Parameters
    ----------
    filename : str
        The filename of the excel file. This file contains every class with the days as the sheet names.
    class_pattern : str
        The class to get the complete time table for. E.g. 'EL 3'

    Returns
    -------
    pandas.DataFrame
        The complete time table for the given class.
    """
    daily_tables = _get_all_daily_tables(filename, class_pattern)

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    for key, value in daily_tables.items():
        if key.title() in days:
            columns = value.columns
            break
    else:
        raise ValueError(f"No sheet found for any of the days: {days}")

    final_df = pd.DataFrame(
        columns=columns,
        index=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
    )

    for day, table in daily_tables.items():
        for period, classes in table.items():
            available_classes = classes.dropna()
            if available_classes.any():
                classrooms = classes[classes.notna()].index
                available_classes = [
                    re.sub(r"\s+", " ", c.strip()) for c in available_classes.values
                ]
                available_classes = [
                    f"{c} ({classrooms[i]})" for i, c in enumerate(available_classes)
                ]
                available_classes = "\n".join(available_classes)
                final_df.loc[day, period] = available_classes

    return final_df


def convert_to_24hour(time_str, is_end_time=False):
    """Convert time to 24-hour format considering class schedule rules."""
    hours, minutes = map(int, time_str.split(':'))
    
    # For start times: hours <= 11 are AM, hours >= 12 are PM
    # For end times: all times are PM
    if is_end_time or hours <= 7:
        if hours != 12:  
            hours += 12
    
    return f"{hours:02d}:{minutes:02d}"


def generate_calendar(timetable, start_date, end_date):
    """
    Generate a calendar of class events based on a given timetable within a specified date range.

    Parameters:
        timetable (list): A list of dictionaries representing the timetable data. Each dictionary contains the following keys:
            - day (str): The name of the day.
            - data (list): A list of dictionaries representing the class events for the day. Each dictionary contains the following keys:
                - start (str): The start time of the class in the format 'HH:MM'.
                - end (str): The end time of the class in the format 'HH:MM'.
                - value (str): The name of the class.

        start_date (str): The start date of the calendar in the format 'YYYY-MM-DD'.
        end_date (str): The end date of the calendar in the format 'YYYY-MM-DD'.

    Returns:
        bytes

    This function generates a calendar of class events based on the provided timetable within the specified date range.
    It iterates over the timetable data, checks if the day matches the current date, and adds class events to the calendar.
    The resulting calendar is saved as an ICS file named 'class_schedule.ics'.
    """
    data = timetable

    cal = Calendar()
    cal.add('version', '2.0')  
    cal.add('prodid', '-//Class Schedule Generator//EN')  

    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")
    
    
    current_date = start_date
    while current_date <= end_date:
        day_name = current_date.strftime("%A")
        for day in data:
            if day["day"] == day_name:
                for class_info in day["data"]:
                    if class_info["value"]:  
                        event = Event()
                        
                        start_time_24h = convert_to_24hour(class_info["start"])
                        end_time_24h = convert_to_24hour(class_info["end"], is_end_time=True)
                        
                        start_time = datetime.strptime(start_time_24h, "%H:%M")
                        end_time = datetime.strptime(end_time_24h, "%H:%M")
                        
                        event_start = current_date.replace(
                            hour=start_time.hour, 
                            minute=start_time.minute
                        )
                        event_end = current_date.replace(
                            hour=end_time.hour, 
                            minute=end_time.minute
                        )
                        
                        event.add("summary", class_info["value"].replace("\n", " "))
                        event.add("dtstart", event_start)
                        event.add("dtend", event_end)
                        event.add('dtstamp', datetime.now())
                        
                        cal.add_component(event)
        current_date += timedelta(days=1)

    with open("class_schedule.ics", "wb") as f:
        f.write(cal.to_ical())
    return cal.to_ical()


def get_all_classes(filename: str) -> List[str]:
    """Extract all class patterns from the workbook."""
    try:
        # Get the path to the Excel file
        UPLOAD_DIR = 'api/drafts'
            
        file_path = os.path.join(UPLOAD_DIR, f"{filename}.xlsx")
        
        # Read the Excel file
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Set to store unique class patterns
        class_patterns = set()
        
        # Department codes from your constants
        dept_codes = [
            'CE', 'MN', 'MC', 'EL', 'GM', 'SD', 'CY', 'PE', 
            'RP', 'GL', 'ME', 'RN', 'IS', 'CH', 'MA', 'ES', 
            'LT', 'LA', 'SP', 'EC'
        ]
        
        # Go through each sheet (day)
        for sheet in wb.worksheets:
            # Scan all cells for class patterns
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip()
                        
                        # Split by commas and process each part
                        parts = cell_value.split(',')
                        for part in parts:
                            part = part.strip()
                            
                            # Handle forward slash pattern (CE/EE 1)
                            slash_pattern = re.search(r'([A-Z]{2,3}(?:/[A-Z]{2,3})+)\s+(\d[A-Z]?)', part)
                            if slash_pattern:
                                depts = slash_pattern.group(1).split('/')
                                year = slash_pattern.group(2)
                                for dept in depts:
                                    if dept in dept_codes:
                                        class_patterns.add(f"{dept} {year}")
                                continue
                            
                            # Match department codes
                            for dept in dept_codes:
                                # Pattern 1: DEPT YEAR[SECTION] (CE 1A, CE 1)
                                matches = re.finditer(
                                    rf'{dept}\s+(\d[A-Z]?)',
                                    part
                                )
                                for match in matches:
                                    class_patterns.add(f"{dept} {match.group(1)}")
                                
                                # Pattern 2: YEAR[SECTION] DEPT (1A CE, 1 CE)
                                matches = re.finditer(
                                    rf'(\d[A-Z]?)\s+{dept}',
                                    part
                                )
                                for match in matches:
                                    class_patterns.add(f"{dept} {match.group(1)}")
                                
                                # Pattern 3: Multiple departments (CE, EE 1)
                                if ',' in part and any(d in part for d in dept_codes):
                                    year_match = re.search(r'\s(\d[A-Z]?)(?:\s|$)', part)
                                    if year_match:
                                        year = year_match.group(1)
                                        dept_matches = re.finditer(r'([A-Z]{2,3})(?:,|\s|$)', part)
                                        for dept_match in dept_matches:
                                            class_patterns.add(f"{dept_match.group(1)} {year}")
        
        # Sort the patterns for consistent output
        return sorted(list(class_patterns))
        
    except Exception as e:
        raise Exception(f"Error reading workbook: {str(e)}")