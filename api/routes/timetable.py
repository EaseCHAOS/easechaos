import os
import logging

from fastapi import APIRouter
from fastapi.responses import FileResponse
from pydantic import BaseModel
from api.extract.extract_table import get_time_table, generate_calendar
import json
from pathlib import Path
import hashlib
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
from fastapi.responses import StreamingResponse

from api.config.redis_config import (
    get_table_from_cache,
    add_table_to_cache,
)


# Find the path of the drafts
current_script_path = Path(__file__)
project_root_path = current_script_path.parents[1]
DRAFTS_FOLDER = project_root_path / "drafts"

router = APIRouter()


class TimeTableRequest(BaseModel):
    """
    Represents a request for a timetable.

    Attributes:
    - filename (str): The name of the file for the timetable.
    - class_pattern (str): The pattern for the class.
    """

    filename: str
    class_pattern: str   

def get_json_table(request: TimeTableRequest):
    """
    A function to get the time table in JSON format.

    Parameters:
    - request: TimeTableRequest - the request object containing the filename and class pattern

    Returns:
    - dict: a dictionary containing the table in JSON format
    """
    table = get_table_from_cache(request.filename, request.class_pattern)

    if table is None:
        filename = os.path.join(DRAFTS_FOLDER, request.filename)
        table = get_time_table(filename, request.class_pattern).to_json(
            orient="records"
        )
        add_table_to_cache(table, request.filename, request.class_pattern)

    return json.loads(table)


# Set up logging
logging.basicConfig(level=logging.ERROR)
logger = logging.getLogger(__name__)

# Track unique problematic entries

def convert_to_24hour(time_str: str, previous_was_pm: bool = False) -> str:
    """Convert time to 24-hour format based on class schedule rules."""
    if not time_str or not time_str.strip():
        raise ValueError("Time string cannot be empty")
        
    try:
        hours, minutes = map(int, time_str.strip().split(':'))
    except ValueError as e:
        raise e
    
    if not previous_was_pm:
        if 7 <= hours <= 11:
            return f"{hours}:{minutes:02d}"
        elif hours == 12:
            return f"12:{minutes:02d}"
        else:
            return f"{hours + 12}:{minutes:02d}"
    else:
        if hours == 12:
            return f"12:{minutes:02d}"
        elif hours <= 7:
            return f"{hours + 12}:{minutes:02d}"
        return f"{hours}:{minutes:02d}"


@router.post("/get_time_table")
async def get_time_table_endpoint(request: TimeTableRequest):
    """
    Endpoint for generating a parsed json time table

    Parameters:
    - request (TimeTableRequest): The request object containing the `filename` and `class_pattern`.

    Returns:
    - JSON: Parsed data from the `get_json_table` function that contains the time table cutting across days and time slots.
        It covers merged durations of lectures exceeding one hour as well.
    """
    table = get_table_from_cache(request.filename, request.class_pattern)
    
    # Generate a hash based on the file content and request parameters
    file_path = os.path.join(DRAFTS_FOLDER, f"{request.filename}.xlsx")
    with open(file_path, "rb") as f:
        content_hash = hashlib.md5(f.read()).hexdigest()
    
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    json_data = get_json_table(request)
    
    table_data = []
    for index, day in enumerate(json_data):
        day_data = []
        current_slot = None
        previous_was_pm = False
        
        for key, value in day.items():
            if not key or not isinstance(key, str):
                continue
                    
            time_parts = key.split("-")
            if len(time_parts) < 2:
                continue
                    
            start = time_parts[0].strip()
            end = time_parts[-1].strip()
                
            if not start or not end:
                continue
                
            try:
                start_24h = convert_to_24hour(start)
                start_hour = int(start_24h.split(':')[0])
                is_pm = start_hour >= 12
                end_24h = convert_to_24hour(end, previous_was_pm)
                    
                if current_slot and current_slot["value"] == value and current_slot["end"] == start_24h:
                    current_slot["end"] = end_24h
                else:
                    if current_slot:
                        day_data.append(current_slot)
                    current_slot = {"start": start_24h, "end": end_24h, "value": value}
                    
                previous_was_pm = is_pm
            except ValueError as e:
                raise e
                    
        if current_slot:
            day_data.append(current_slot)
            
        table_data.append({"day": days[index], "data": day_data})


    return {
        "data": table_data,
        "version": content_hash
    }


@router.post("/download")
async def download_time_table_endpoint(request: TimeTableRequest):
    """
    Endpoint for downloading a time table as an Excel file.

    Parameters:
    - request (TimeTableRequest): The request object containing the filename and class pattern.

    Returns:
    - FileResponse: The Excel file containing the time table.

    Description:
    This function is an endpoint for downloading a time table as an Excel file.
    It takes a `TimeTableRequest` object as a parameter, which contains the filename and class pattern.
    The function first checks if the time table is already cached. If it is, it retrieves the cached table.
    Otherwise, it generates the time table by calling the `get_time_table` function and adds it to the cache.
    The function then converts the time table into a Pandas DataFrame and creates an Excel file using the `openpyxl` library.
    It iterates over the columns and rows of the DataFrame and writes the values to the Excel worksheet.
    Finally, it saves the Excel file to a buffer and returns it as a `FileResponse` object with the appropriate media type.

    Note:
    - The `TimeTableRequest` class should have the following attributes:
        - filename (str): The name of the file for the time table.
        - class_pattern (str): The pattern for the class.
    - The `get_table_from_cache` and `add_table_to_cache` functions should be implemented elsewhere in the codebase.

    ```
    """
    filename = os.path.join(DRAFTS_FOLDER, request.filename)
    table = get_table_from_cache(request.filename, request.class_pattern)

    if table is None:
        table = get_time_table(filename, request.class_pattern).to_json(
            orient="records"
        )
        add_table_to_cache(
            table=table, filename=request.filename, class_pattern=request.class_pattern
        )
    
    df = pd.DataFrame(json.loads(table))
    buffer = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active

    for col_index, col_name in enumerate(df.columns, start=1):
        worksheet.cell(row=1, column=col_index, value=col_name)

    for row_index, row in enumerate(df.itertuples(), start=2):
        for col_index, value in enumerate(row[1:], start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    workbook.save(buffer)

    excel_content = buffer.getvalue()
    return FileResponse(
        excel_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/calendar_file")
async def calendar_file_endpoint(request: TimeTableRequest):
    """
    Endpoint for generating a calendar file.

    Args:
        request (TimeTableRequest): The request object containing the `filename` and `class_pattern`.

    Returns:
        StreamingResponse: A streaming response containing the calendar file. The file is downloaded with the name "class_schedule.ics".

    Description:
        This function is an endpoint for generating a calendar file. It takes a `TimeTableRequest`
        object as a parameter, which contains the `filename` and `class_pattern`.
        The function first calls the `get_time_table_endpoint` function to get the time table.
        It then generates a calendar file using the `generate_calendar`
        function with the provided time table, start date, and end date.
        The generated calendar file is stored in a `BytesIO` object.
        The function sets the file pointer to the beginning of the file and creates a streaming response with the calendar file.
        The response is set to have the media type "text/calendar".
        The function sets the "Content-Disposition" header of the response to "attachment;
        filename=class_schedule.ics" to indicate that the file should be downloaded.
        Finally, the function returns the streaming response.
    """

    timetable = await get_time_table_endpoint(request)
    start_date = "2023-01-01"
    end_date = "2023-02-01"
    cal = generate_calendar(
        timetable=timetable, start_date=start_date, end_date=end_date
    )
    cal_file = BytesIO(cal)
    cal_file.seek(0)

    # Return as a downloadable file response
    response = StreamingResponse(cal_file, media_type="text/calendar")
    response.headers["Content-Disposition"] = "attachment; filename=class_schedule.ics"
    return response
